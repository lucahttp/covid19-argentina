import time
import os.path
import datetime
from threading import Thread
import pandas as pd
import sqlite3
from flask import Flask, request, render_template, json, jsonify, send_from_directory, g
import os
from datetime import timedelta
import requests
#from multiprocessing import Process


# check if db exist

# check if the expiration date has passed

# do
# delete csv
# delete db
# download csv
# create db

# pass

# do
# download csv
# create db


class CovidData:
    def __init__(self, url):
        self.url = url
        self.state = None
        self.LastUpdate = None

        # Files
        #self.filename = "covid_3.sql"
        self.csvfilepath = './data/casoscovid19.csv'
        self.dbfilepath = './data/casoscovid19.db'
        self.mytable = 'mydb'

        self.csvfilereport = "report.csv"

        self.directory = './data/'
        # SQLite
        self.conn = None

        # Data
        self.countAllCases = None

        # Flags
        self.flag_downloading = False

        self.flag_download_current = 0
        self.flag_download_total = 0
        self.flag_download_percent = 0

        self.flag_creating = False

    def getCountAllCases(self):
        return self.countAllCases

    def getStatus(self):
        return self.state

    def setStatus(self, newStatus):
        self.state = newStatus
        print("new status setted")

        # get file date

    def getLastUpdateOfFile(self, file):

        print()
        print("file : "+file)

        created = os.path.getctime(file)
        return created

    # add time to a specific date to create the due date

    def setPassDay(self, mydate):

        #print("incoming date " + str(mydate))

        theNextDay = mydate + timedelta(days=1)
        theNextDay = theNextDay.replace(
            hour=20, minute=00, second=00, microsecond=00)

        #print("outgoing date " + str(theNextDay))

        from tabulate import tabulate
        #print(tabulate([['2020-08-14 16:36:08','2020-08-15 20:00:00','2020-08-14 19:15:00'], []], headers=["File Date", "File Expiration Date","Current Date"]))
        print()
        print(tabulate([[str(mydate.replace(microsecond=00)), str(theNextDay), str(datetime.datetime.now(
        ).replace(microsecond=00))], []], headers=["File Date", "File Expiration Date", "Current Date"]))

        return theNextDay

    def CheckDueDate(self, mydate, when):

        # print(mydate)
        #print("the date of the file date is: " + str(datetime.datetime.fromtimestamp(mydate)))

        # from tabulate import tabulate
        # print(tabulate([['2020-08-14 16:36:08','2020-08-15 20:00:00','2020-08-14 19:15:00'], []], headers=["File Date", "File Expiration Date","Current Date"]))

        myInputDate = datetime.datetime.fromtimestamp(mydate)

        myOut = None

        if self.setPassDay(myInputDate) < datetime.datetime.now():
            #print("es mas grande")
            #print("ya vencio"+" - es mas grande")
            myOut = True
            pass
        else:
            #print("es mas chico")
            #print("vencido = "+str(covidargentina.CheckData()))
            #print("todavia no vencio"+" - es mas chico")
            myOut = False
            pass

        print("vencido = "+str(myOut))

        return myOut

    # Downloader
    # https://medium.com/@petehouston/download-files-with-progress-in-python-96f14f6417a2

    # import requests
    def download(self, url, filefolder):
        import wget
        print("Download started")
        #myfile = requests.get(url, allow_redirects=True)

        def bar_custom(current, total, width=80):
            self.flag_download_current = current
            self.flag_download_total = total
            self.flag_download_percent = current / total * 100
            print("Downloading: %d%% [%d / %d] bytes" %
                  (current / total * 100, current, total))

        #wget.download(url, bar=bar_custom)
        wget.download(url, filefolder, bar=bar_custom)
        #open(filefolder, 'wb').write(myfile.content)
        print("Download finished")
        pass

    def downloadCSV(self):
        # Downloader
        # https://stackoverflow.com/questions/22676/how-do-i-download-a-file-over-http-using-python
        print("Downloader")
        import os

        if self.flag_downloading == False:
            self.flag_downloading = True

            if not os.path.exists(self.directory):
                os.makedirs(self.directory)
                self.download(self.url, self.csvfilepath)
                pass
            else:
                self.download(self.url, self.csvfilepath)
                pass

            self.flag_downloading = False
            pass
        else:
            print("Is already downloading")
            pass

    def deleteAndRemove(self, filetopath):
        import os
        #filePath = '/home/somedir/Documents/python/logs';
        # As file at filePath is deleted now, so we should check if file exists or not not before deleting them
        try:
            os.remove(filetopath)
            pass
        except FileNotFoundError:
            print("Can not delete the file as it doesn't exists")
            pass
        """
        if os.path.exists(filetopath):
            os.remove(filetopath)
        else:
            print("Can not delete the file as it doesn't exists")
        pass
        """

    def createDB(self, csvpath, dbpath):
        # read the CSV
        #df = pd.read_csv('./data/casoscovid19.csv')
        df = pd.read_csv(csvpath)
        # connect to a database
        #conn = sqlite3.connect("casoscovid19.db")
        self.conn = sqlite3.connect(dbpath)
        # if the db does not exist, this creates a Any_Database_Name.db file in the current directory
        # store your table in the database:

        cursor = self.conn.cursor()
        # Doping EMPLOYEE table if already exists
        cursor.execute("DROP TABLE IF EXISTS mydb;")
        #print("Table dropped... ")
        # Commit your changes in the database
        self.conn.commit()
        # Closing the connection
        # conn.close()
        # https://www.tutorialspoint.com/python_sqlite/python_sqlite_drop_table.htm

        #gg = pd.read_sql('DROP TABLE IF EXISTS test_0;', conn)
        #pd.read_sql_query('DROP TABLE IF EXISTS test_0;', conn)
        df.to_sql("mydb", self.conn)
        pass

    def onlyConnectToDB(self):
        print("we only create the connection to the DB")
        # connect to a database
        self.conn = sqlite3.connect(self.dbfilepath)
        pass

    def CheckData(self):
        try:
            result = self.CheckDueDate(self.getLastUpdateOfFile(
                self.csvfilepath), "1 day at 8 pm")
            return result
        except FileNotFoundError:
            print("No funciona")
            self.deleteAndRemove(self.csvfilepath)
            self.deleteAndRemove(self.dbfilepath)
            self.downloadCSV()
            self.createDB(self.csvfilepath, self.dbfilepath)
            """
            self.createDB(self.csvfilepath, self.dbfilepath)
            self.conn = sqlite3.connect(self.dbfilepath)
            
            """
            pass

    def checkAllsGood(self):
        if os.path.isfile(self.csvfilepath):
            print("File exist")

            # test if file pass the due date
            if self.CheckDueDate(self.getLastUpdateOfFile(self.csvfilepath), "1 day at 8 pm"):
                print("pass the due date")

                if covidargentina.flag_downloading == True:
                    print()
                    pass
                else:
                    self.deleteAndRemove(self.csvfilepath)
                    self.deleteAndRemove(self.dbfilepath)
                    self.downloadCSV()
                    self.createDB(self.csvfilepath, self.dbfilepath)
                    self.conn = sqlite3.connect(self.dbfilepath)
                    pass

                pass
            else:
                print("not pass the due date")
                print("all okay")
                print()
                print("we only create the connection to the DB")
                # connect to a database

                filenamepath = self.csvfilepath
                if os.path.isfile(filenamepath):
                    print("File exist")
                    #filenamecontent = open(filenamepath)
                    #dataso = filenamecontent.read()
                else:
                    #dataso = {'thread_name': str(thread.name),'started': True, 'Status': 'please wait'}
                    print("File not exist")
                    self.createDB(self.csvfilepath, self.dbfilepath)

                self.conn = sqlite3.connect(self.dbfilepath)
                pass
        else:
            print("File not exist")
            self.downloadCSV()
            self.createDB(self.csvfilepath, self.dbfilepath)
            self.conn = sqlite3.connect(self.dbfilepath)
            pass
        pass

    def superRefresh(self):
        self.deleteAndRemove(self.csvfilepath)
        self.deleteAndRemove(self.dbfilepath)
        self.downloadCSV()
        self.createDB(self.csvfilepath, self.dbfilepath)
        self.conn = sqlite3.connect(self.dbfilepath)
        pass
    #print(CheckDueDate(getLastUpdateOfFile("covid_3.sql"),"1 day at 8 pm"))

    # check if db exist

        # check if the expiration date has passed

        # do
        # delete csv
        # delete db
        # download csv
        # create db

        # pass

        # do
        # download csv
        # create db

    # test if file exist
    # https://linuxize.com/post/python-check-if-file-exists/
    def makeReport(self, gg, report_name):
        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        print(type(gg))
        gg.to_csv(report_name+".csv", encoding='latin1', index=False)
        #json.dumps(parsed, indent=4, ensure_ascii=False)
        # gg.set_index(list(gg)[0])
        gg = gg.set_index(gg.columns[0])
        gg.set_index(gg.columns.tolist()[0])
        #json.dumps(parsed, indent=4)

        result = gg.to_json(orient="index")
        parsed = json.loads(result)
        resultado = gg.to_json(orient="index")
        gg.to_json(report_name+".json", orient='table',
                   force_ascii=False, indent=4)

        # Works
        #out = json.dumps(parsed, indent=4, ensure_ascii=False)
        #gg.to_csv('report.csv', encoding='utf-8', index=False)

        #result = gg.to_json(orient="index")

        parsed = json.loads(result)

        json.dumps(parsed, indent=4)
        # print(resultado)
        return resultado

    def workWithOnlyCSV(self):
        try:
            gg = self.CheckDueDate(self.getLastUpdateOfFile(
                self.csvfilepath), "1 day at 8 pm")
            if gg:
                print("pass the due date")
                self.deleteAndRemove(self.csvfilepath)
                self.downloadCSV()
                pass
            else:
                #print("not pass the due date")
                # print("okay")
                # print()
                #print("we only create the connection to the DB")
                pass
            pass
        except FileNotFoundError:
            gg = False
            self.downloadCSV()
            pass

    def totalCasosConfirmados(self):
        self.workWithOnlyCSV()
        # https://www.geeksforgeeks.org/python-filtering-data-with-pandas-query-method/
        # importing pandas package
        # import pandas as pd

        # making data frame from csv file
        data = pd.read_csv("./data/casoscovid19.csv")

        # replacing blank spaces with '_'
        data.columns = [column.replace(" ", "_") for column in data.columns]

        # filtering with query method
        #print(len(data.query('clasificacion_resumen =="Confirmado"', inplace = True)))
        return len(data.query('clasificacion_resumen =="Confirmado"'))

    """
    SELECT residencia_departamento_nombre  AS Residencia,residencia_provincia_nombre  AS Provincia,
        count(*) AS Total,
        sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
        sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
        sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
    FROM self.mytable
    GROUP BY residencia_departamento_nombre
    ORDER BY "residencia_provincia_id" ASC,"residencia_departamento_nombre" ASC,"residencia_departamento_nombre" ASC;
    """

    def moredata(self):
        self.workWithOnlyCSV()
        # https://www.geeksforgeeks.org/python-filtering-data-with-pandas-query-method/
        # importing pandas package
        # import pandas as pd

        # making data frame from csv file
        data = pd.read_csv("./data/casoscovid19.csv")

        # replacing blank spaces with '_'
        data.columns = [column.replace(" ", "_") for column in data.columns]

        # filtering with query method
        #print(len(data.query('clasificacion_resumen =="Confirmado"', inplace = True)))
        self.countAllCases = len(data.query(
            'clasificacion_resumen =="Confirmado"'))
        pass

    def fullreportSimple(self):
        self.checkAllsGood()
        sql_string = """
        SELECT residencia_departamento_nombre  AS Residencia,residencia_provincia_nombre  AS Provincia,
            count(*) AS Total,
            sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
            sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
            sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM mydb
        GROUP BY residencia_departamento_nombre
        ORDER BY "residencia_provincia_id" ASC,"residencia_departamento_nombre" ASC,"residencia_departamento_nombre" ASC;
        """
        gg = pd.read_sql(sql_string, self.conn)
        print()

        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        print(type(gg))
        gg.to_csv("report.csv", encoding='latin1', index=False)
        #json.dumps(parsed, indent=4, ensure_ascii=False)
        # gg.set_index(list(gg)[0])
        gg = gg.set_index(gg.columns[0])
        gg.set_index(gg.columns.tolist()[0])
        #json.dumps(parsed, indent=4)

        result = gg.to_json(orient="index")
        parsed = json.loads(result)
        resultado = gg.to_json(orient="index")
        gg.to_json("report.json", orient='table', force_ascii=False, indent=4)

        # Works
        #out = json.dumps(parsed, indent=4, ensure_ascii=False)
        #gg.to_csv('report.csv', encoding='utf-8', index=False)

        #result = gg.to_json(orient="index")

        parsed = json.loads(result)

        json.dumps(parsed, indent=4)
        # print(resultado)
        return resultado

    def fullreport(self):
        covidargentina.checkAllsGood()
        # https://stackoverflow.com/questions/4899832/sqlite-function-to-format-numbers-with-leading-zeroes
        # https://tiebing.blogspot.com/2011/07/sqlite-3-string-to-integer-conversion.html
        """
        CAST(substr('00'||residencia_provincia_id,-2) || substr('000'||residencia_departamento_id,-3) as integer) AS ID,
        """
        sql_string = """
        SELECT residencia_departamento_nombre  AS "Departamento Residencia",residencia_provincia_nombre  AS "Provincia Residencia", 
        substr('00'||residencia_provincia_id,-2) || substr('000'||residencia_departamento_id,-3) AS ID,
        substr('000'||residencia_departamento_id,-3) AS "ID Departamento",
        substr('00'||residencia_provincia_id,-2) AS "ID Provincia",
        count(*) AS "Total Test",
        sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
        sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
        (sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end)-sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end))Activos,
        sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM mydb
        GROUP BY residencia_departamento_nombre
        ORDER BY "residencia_provincia_id" ASC,"residencia_departamento_nombre" ASC,"residencia_departamento_nombre" ASC;
        """
        gg = pd.read_sql(sql_string, self.conn)
        print()

        return self.makeReport(gg, "fullreport")

    def datostableargentina(self):
        # table argentina
        # Provincia	Casos Confirmados Totales	Recuperados	Fallecidos
        sql_string = '''SELECT count(*) AS Total,
            sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
            sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
            sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM '''+self.mytable+''';
        '''
        gg = pd.read_sql(sql_string, self.conn)
        print()

        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        # print(gg)
        # print(type(gg))
        # return gg.to_csv(encoding='latin1', index=False)
        self.makeReport(gg, "argentinareport")
        return gg

    def datoshistoricos(self):
        # table argentina
        # Provincia	Casos Confirmados Totales	Recuperados	Fallecidos
        sql_string = '''SELECT fecha_apertura  AS "Fecha",
            count(*) AS "Cantidad de test",
            sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
            sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
            (sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end)-sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end))Activos,
            sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM '''+self.mytable+'''
        GROUP BY fecha_apertura
        ORDER BY "fecha_apertura" DESC;
        '''
        gg = pd.read_sql(sql_string, self.conn)
        print()

        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        # print(gg)
        # print(type(gg))
        # return gg.to_csv(encoding='latin1', index=False)
        return self.makeReport(gg, "historicosreport")
        return gg



    def datostableprovincias(self):
        # table provincias
        # Provincia	Casos Confirmados Totales	Recuperados	Fallecidos
        sql_string = '''SELECT residencia_provincia_nombre  AS Provincia,
            count(*) AS Total,
            sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
            sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
            sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM '''+self.mytable+'''
        GROUP BY residencia_provincia_nombre;
        '''
        gg = pd.read_sql(sql_string, self.conn)
        print()

        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        # print(gg)
        # print(type(gg))

        # return gg.to_csv(encoding='latin1', index=False)
        self.makeReport(gg, "provinciasreport")
        return gg

    def reportcsv(self):
        sql_string = """
        SELECT residencia_departamento_nombre  AS Residencia,residencia_provincia_nombre  AS Provincia,
            count(*) AS Total,
            sum(case when clasificacion_resumen="Confirmado" then 1 else 0 end) AS Confirmados,
            sum(case when clasificacion LIKE "%No Activo%" then 1 else 0 end) AS Recuperados,
            sum(case when clasificacion = "Caso confirmado - Fallecido" then 1 else 0 end) AS Fallecidos
        FROM mydb
        GROUP BY residencia_departamento_nombre
        ORDER BY "residencia_provincia_id" ASC,"residencia_departamento_nombre" ASC,"residencia_departamento_nombre" ASC;
        """
        gg = pd.read_sql(sql_string, self.conn)
        print()

        import json
        # print(type(gg))
        #gg = gg.set_index(0)
        print(gg)

        return gg.to_csv(encoding='latin1', index=False)



    ################################################################################
    # Tool Functions                        function to use cross program

    def GetRandom(self):
        #https://docs.python.org/3/library/random.html
        import random
        random_number = random.randrange(0, 10000, 1)
        self.printandlog(random_number)
        #excelFile = "tempFile-{}.xlsx".format(str(current_time))
        #excelFile = "tempFile-{}.xlsx".format(random_number)
        return random_number

    def getTime(self):
        import datetime
        x = datetime.datetime.now()
        xs = x.strftime("%X %x")
        datestring = str(xs)
        datestring = datestring.replace(":","-")
        datestring = datestring.replace("/", "-")
        datestring = datestring.replace(" ", "--")
        print(datestring)
        return datestring



    ################################################################################
    #LOG                        function to log printandlogs

    

    def printandlog(self,text):
        import logging
        logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s %(message)s')
        print(text)
        logging.warning(text)
        pass




    def CreateExample(self):
        from time import sleep
        import datetime

        # C:\Python38\python.exe -m pip install --upgrade pip
        # C:\Python38\python.exe -m pip install openpyxl
        # C:\Python38\python.exe -m pip install pytest
        import pytest
        import time
        import json

        #EXCEL
        from openpyxl import load_workbook
        import os
        import sys
        import openpyxl #Connect the library
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill#Connect cell styles
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Font, Fill#Connect styles for text
        from openpyxl.styles import colors#Connect colors for text and cells
        from openpyxl.utils import get_column_letter
        import datetime
        import string







        ##########################################################
        ##  Chapter One Creating the excel to fill data
        ##########################################################


        #https://stackabuse.com/the-python-tempfile-module/


        tempfilename = "~tempfile.xlsx"


        datestring = self.getTime()
        #with io.open((path.replace("xlsx", "{}.json")).replace("xls", "{}.json").format(number,number),'w',encoding='utf8') as f:

        ##tempfilename = "~tempfile{}.xlsx".format(datestring)

        
        #delete file if exist feature

        if os.path.exists(tempfilename):
            os.remove(tempfilename)
            self.printandlog("The file was deleted")
        else:
            self.printandlog("The file does not exist")
        


        #Styles
        tabColor_text = Font(size=11, underline='none', color = 'ffffffff', bold=False, italic=False) #what color = colors.RED — color prescribed in styles
        tabColor_cell = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')#This code allows you to do design color cells


        wb = Workbook()
        page = wb.active
        #Set the name of the workbook
        page.title = 'TO_SERVICE-NOW'

        #Insert the example data to the example excel    
        #TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
        TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
        page.append(TabTitles) # write the TabTitles to the first line

        #Insert the example data to the example excel
        #companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]
        companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]


        page.append(companies)

        # Apply styles to the example sheet
        for i in range (1, len(TabTitles)+1): 
            self.printandlog(TabTitles[i - 1])

            cell_obj = page.cell(row = 1, column = i)
            #cell_obj.value = TabTitles[i - 1]
            self.printandlog(len(cell_obj.value))
            current_col = string.ascii_uppercase[i - 1]

            page.column_dimensions[current_col].width = len(cell_obj.value) * 1.5
            cell_obj.fill = tabColor_cell 
            cell_obj.font = tabColor_text
            self.printandlog("") 

        # excel var utilities
        max_row = page.max_row
        max_col = page.max_column
        factor_charactersize = 1.18

        ##testing
        #printandlog("testing")

        #var that store the max width of a cell in a column
        max_long_per_column = []

        #printandlog("")

        #to fix the size of the cells in the temporal excel

        for x in range(0, max_col):
            current_col = string.ascii_uppercase[x]
            #numbers[x] = arr.array([0])
            #max_long_per_column[x].append(0)
            max_long_per_column.append(0)
            for y in range(1, max_row + 1):
                #save cell value
                cell_obj = page.cell(row = y, column = x + 1)
                self.printandlog(cell_obj)
                self.printandlog(type(cell_obj.value).__name__)

                if type(cell_obj.value).__name__ != 'NoneType':
                    self.printandlog(len(cell_obj.value))

                    #
                    if max_long_per_column[x] < len(cell_obj.value):
                        self.printandlog("encontro uno mas grande")
                        max_long_per_column[x] = len(cell_obj.value)
                        self.printandlog("el mas grande: " + str(max_long_per_column[x]))

                        page.column_dimensions[current_col].width = max_long_per_column[x] * factor_charactersize
                        pass
                    else:
                        self.printandlog("encontro uno mas chico")
                        pass
                    pass
                else:
                    self.printandlog("WTF")
                    pass
            self.printandlog("Termino uno calumna --------------------------------------------")
            #check

            #page.column_dimensions[current_col].width = max_long_per_column[x] * 1.5
            
        #workbook_name = 'sample.xlsx'
        #workbook_name_temp = ''

        #excelFile = "{}\\~tempFile-{}.xlsx".format(sys.path[0], str(datetime.datetime.now()))
        #excelFile = "{}\\~tempfile.xlsx".format(sys.path[0])





        #printandlog("Goto Sleep")
        #sleep(50)

        """
        for info in companies:
            page.append(info)
        """
        #wb.save(filename = workbook_name)

        wb.save(filename = tempfilename)


        #stuff_in_string = 'start excel.exe "{}"'.format(workbook_name)

        #stuff_in_string = 'start excel.exe "{}"'.format(excelFile)
        #stuff_in_string = 'start excel "{}"'.format(excelFile)

        #os.system(stuff_in_string)
        pass



    def CreateExcelFormated(self,mydataframe):
        from time import sleep
        import datetime

        # C:\Python38\python.exe -m pip install --upgrade pip
        # C:\Python38\python.exe -m pip install openpyxl
        # C:\Python38\python.exe -m pip install pytest
        import pytest
        import time
        import json

        #EXCEL
        from openpyxl import load_workbook
        import os
        import sys
        import openpyxl.utils.dataframe as oput #Connect the library
        import openpyxl #Connect the library
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill#Connect cell styles
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Font, Fill#Connect styles for text
        from openpyxl.styles import colors#Connect colors for text and cells
        from openpyxl.utils import get_column_letter
        import datetime
        import string


        import numpy




        ##########################################################
        ##  Chapter One Creating the excel to fill data
        ##########################################################


        #https://stackabuse.com/the-python-tempfile-module/


        tempfilename = "~tempfile.xlsx"


        datestring = self.getTime()
        #with io.open((path.replace("xlsx", "{}.json")).replace("xls", "{}.json").format(number,number),'w',encoding='utf8') as f:

        ##tempfilename = "~tempfile{}.xlsx".format(datestring)

        
        #delete file if exist feature

        if os.path.exists(tempfilename):
            os.remove(tempfilename)
            self.printandlog("The file was deleted")
        else:
            self.printandlog("The file does not exist")
        


        #Styles
        tabColor_text = Font(size=11, underline='none', color = 'ffffffff', bold=False, italic=False) #what color = colors.RED — color prescribed in styles
        tabColor_cell = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')#This code allows you to do design color cells


        wb = Workbook()
        page = wb.active
        #Set the name of the workbook
        page.title = 'TO_SERVICE-NOW'


        # mydataframe
        # Create a Pandas Excel writer using XlsxWriter as the engine. 
        

        content = pd.DataFrame(oput.dataframe_to_rows(mydataframe, index=True, header=True)).to_numpy()
        #Insert the example data to the example excel    
        #TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
        TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
        #page.append(TabTitles) # write the TabTitles to the first line
        
        #Insert the example data to the example excel
        #companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]
        companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]

        #page.append(companies)
        print(" - ")
        print(content.to_numpy())
        print(" - ")
        page.append(content)

        # Apply styles to the example sheet
        for i in range (1, len(TabTitles)+1): 
            self.printandlog(TabTitles[i - 1])

            cell_obj = page.cell(row = 1, column = i)
            #cell_obj.value = TabTitles[i - 1]
            self.printandlog(len(cell_obj.value))
            current_col = string.ascii_uppercase[i - 1]

            page.column_dimensions[current_col].width = len(cell_obj.value) * 1.5
            cell_obj.fill = tabColor_cell 
            cell_obj.font = tabColor_text
            self.printandlog("") 

        # excel var utilities
        max_row = page.max_row
        max_col = page.max_column
        factor_charactersize = 1.18

        ##testing
        #printandlog("testing")

        #var that store the max width of a cell in a column
        max_long_per_column = []

        #printandlog("")

        #to fix the size of the cells in the temporal excel

        for x in range(0, max_col):
            current_col = string.ascii_uppercase[x]
            #numbers[x] = arr.array([0])
            #max_long_per_column[x].append(0)
            max_long_per_column.append(0)
            for y in range(1, max_row + 1):
                #save cell value
                cell_obj = page.cell(row = y, column = x + 1)
                self.printandlog(cell_obj)
                self.printandlog(type(cell_obj.value).__name__)

                if type(cell_obj.value).__name__ != 'NoneType':
                    self.printandlog(len(cell_obj.value))

                    #
                    if max_long_per_column[x] < len(cell_obj.value):
                        self.printandlog("encontro uno mas grande")
                        max_long_per_column[x] = len(cell_obj.value)
                        self.printandlog("el mas grande: " + str(max_long_per_column[x]))

                        page.column_dimensions[current_col].width = max_long_per_column[x] * factor_charactersize
                        pass
                    else:
                        self.printandlog("encontro uno mas chico")
                        pass
                    pass
                else:
                    self.printandlog("WTF")
                    pass
            self.printandlog("Termino uno calumna --------------------------------------------")
            #check

            #page.column_dimensions[current_col].width = max_long_per_column[x] * 1.5
            
        #workbook_name = 'sample.xlsx'
        #workbook_name_temp = ''

        #excelFile = "{}\\~tempFile-{}.xlsx".format(sys.path[0], str(datetime.datetime.now()))
        #excelFile = "{}\\~tempfile.xlsx".format(sys.path[0])





        #printandlog("Goto Sleep")
        #sleep(50)

        """
        for info in companies:
            page.append(info)
        """
        #wb.save(filename = workbook_name)

        wb.save(filename = tempfilename)


        #stuff_in_string = 'start excel.exe "{}"'.format(workbook_name)

        #stuff_in_string = 'start excel.exe "{}"'.format(excelFile)
        #stuff_in_string = 'start excel "{}"'.format(excelFile)

        #os.system(stuff_in_string)
        pass




# some JSON:
x = '{ "name":"John", "age":30, "city":"New York"}'


url = "https://sisa.msal.gov.ar/datos/descargas/covid-19/files/Covid19Casos.csv"
covidargentina = CovidData(url)
# covidargentina.getStatus()

#print("vencido = "+str(covidargentina.CheckData()))
# covidargentina.downloadCSV()
# Define some heavy function
# covidargentina.checkAllsGood()
#covidargentina.setStatus("Process started")

# print(covidargentina.getStatus())
# time.sleep(10)


#covidargentina.setStatus("Process finished")

# print(covidargentina.getStatus())


#print("vencido = "+str(covidargentina.CheckData()))

# print()
# print(covidargentina.moredata())

# parse x:
#y = json.loads(x)

# the result is a Python dictionary:


covidargentina.checkAllsGood()
covidargentina.datostableprovincias()
covidargentina.datostableargentina()

# covidargentina.CreateExample()
covidargentina.CreateExcelFormated(covidargentina.datoshistoricos())
print("GG")

# covidargentina.createDB(covidargentina.csvfilepath,covidargentina.dbfilepath)
# datosqwe = covidargentina.fullreport()
# print(datosqwe)


def toMegabyte(number_in_bytes):
    #return round(number_in_bytes/1048576, 2)
    return round(number_in_bytes/1048576, 2)


def make_summary_csv():
    return covidargentina.reportcsv()

# workWithOnlyCSV()


def myfunct():
    print("gg")
    covidargentina.moredata()
    pass


app = Flask(__name__, static_folder='static', static_url_path='')

# Async in flask
# https://stackoverflow.com/questions/31866796/making-an-asynchronous-task-in-flask


@app.route('/test/', methods=['POST', 'GET'])
def asdasd():
    # return Response(mimetype='application/json', status=200)
    data = "Works!!!"
    response = app.response_class(
        response=data,
        # response=data,
        status=200,
        mimetype='text/plain'
    )
    return response


@app.route('/render/<id>', methods=['GET'])
def render_script(id=None):
    # https://smirnov-am.github.io/background-jobs-with-flask/
    thread = Thread(target=myfunct)
    thread.daemon = True
    thread.start()
    # return Response(mimetype='application/json', status=200)
    data = covidargentina.getCountAllCases()
    response = app.response_class(
        response="GG",
        # response=data,
        status=200,
        mimetype='text/plain'
    )
    return response

#data = 'hola : luca'


@app.route("/", defaults={'elque': None, 'como': None, })
@app.route('/report/<elque>/<como>', methods=['POST', 'GET'])
def reportprovincias_func(elque=None, como=None):

    thread = Thread(target=my_multi_report_func, args=(elque, como))
    #thread = Thread(target=my_reportprovincias_func)
    thread.daemon = True
    thread.start()

    if elque == "argentina":
        filenamepath = 'argentinareport'
        pass

    elif elque == "provincias":
        filenamepath = 'provinciasreport'
        pass

    elif elque == "departamentos":
        filenamepath = 'fullreport'
        pass
    elif elque == "historico":
        filenamepath = 'historicosreport'
        pass

    else:
        filenamepath = 'argentinareport'
        pass

    if como == "json":
        extencionfilename = '.json'
        mymimetype = 'application/json'
        pass

    elif como == "csv":
        extencionfilename = '.csv'
        mymimetype = 'text/plain'
        pass
    else:
        extencionfilename = '.json'
        mymimetype = 'application/json'
        pass

    fullfilenamepath = filenamepath+extencionfilename

    if os.path.isfile(fullfilenamepath):
        print("File exist")
        filenamecontent = open(fullfilenamepath)
        dataso = filenamecontent.read()
    else:
        dataso = {'thread_name': str(thread.name),
                  'started': True, 'Status': 'please wait'}
        print("File not exist")

    response = app.response_class(
        response=dataso,
        # response=data,
        status=200,
        mimetype=mymimetype
        # text/plain, text/html, text/css, text/javascript application/json
        # https://developer.mozilla.org/es/docs/Web/HTTP/Basics_of_HTTP/MIME_types
    )
    return response
# Define some heavy function


def my_multi_report_func(posiblefilename, extencionfilename):
    # time.sleep(1)
    print("Download process started")
    covidargentina.checkAllsGood()
    print("Download process finished")

    print("Build DB process started")

    if posiblefilename == "provincias":
        covidargentina.datostableprovincias()
        pass

    elif posiblefilename == "argentina":
        covidargentina.datostableargentina()
        pass

    elif posiblefilename == "departamentos":
        covidargentina.fullreport()
        pass
    elif posiblefilename == "historico":
        covidargentina.datoshistoricos()
        pass
    else:
        pass
    print("Build DB  process finished")
    # setData()
    pass


def my_reportprovincias_func():
    # time.sleep(1)
    print("Download process started")
    covidargentina.checkAllsGood()
    print("Download process finished")

    print("Build DB process started")
    covidargentina.datostableprovincias()
    print("Build DB  process finished")
    # setData()
    pass


@app.route('/report', methods=['POST', 'GET'])
def script():
    thread = Thread(target=my_func)
    thread.daemon = True
    thread.start()

    filenamepath = 'report.json'
    if os.path.isfile(filenamepath):
        print("File exist")
        filenamecontent = open(filenamepath)
        dataso = filenamecontent.read()
    else:
        dataso = {'thread_name': str(thread.name),
                  'started': True, 'Status': 'please wait'}
        print("File not exist")

    response = app.response_class(
        response=dataso,
        # response=data,
        status=200,
        mimetype='application/json'
        # text/plain, text/html, text/css, text/javascript application/json
        # https://developer.mozilla.org/es/docs/Web/HTTP/Basics_of_HTTP/MIME_types
    )
    return response
# Define some heavy function


def my_func():
    # time.sleep(1)
    print("Download process started")
    covidargentina.checkAllsGood()
    print("Download process finished")

    print("Build DB process started")
    covidargentina.fullreport()
    print("Build DB  process finished")
    # setData()
    pass


def make_summary():
    x = '{"name":"luca"}'
    x = '{"casos":'+str(covidargentina.totalCasosConfirmados())+'}'
    data = json.loads(x)
    return data


@app.route('/preview', methods=['GET', 'POST'])
def preview():
    return render_template('table.html')


# https://stackoverflow.com/questions/4239825/static-files-in-flask-robot-txt-sitemap-xml-mod-wsgi


@app.route("/refresh")
def index():
    thread = Thread(target=my_refresh)
    thread.daemon = True
    thread.start()

    data = {'thread_name': str(thread.name),
            'started': True, 'Status': 'please wait'}

    data = jsonify({'thread_name': str(thread.name),
                    'started': True,
                    'Status': 'please wait'})
    response = app.response_class(
        response=data,
        # response=data,
        status=200,
        mimetype='application/json'
        # text/plain, text/html, text/css, text/javascript application/json
        # https://developer.mozilla.org/es/docs/Web/HTTP/Basics_of_HTTP/MIME_types
    )
    return data


def my_refresh():
    # time.sleep(1)
    print("Refresh process started")
    covidargentina.superRefresh()
    # covidargentina.deleteAndRemove(covidargentina.csvfilepath)
    # covidargentina.deleteAndRemove(covidargentina.dbfilepath)
    #print("Download process started")
    # covidargentina.downloadCSV()
    #print("Download process finished")
    # covidargentina.checkAllsGood()

    #print("Build DB process started")
    # covidargentina.fullreport()
    #print("Build DB  process finished")
    print("Refresh process finished")
    # setData()
    pass


@app.route('/preview/csv', methods=['GET', 'POST'])
def catch_all():
    f = open('report.csv')
    return f.read()


@app.route('/', methods=['GET', 'POST'])
def test():
    if covidargentina.flag_downloading == False:
        covidargentina.moredata()
        data = covidargentina.getCountAllCases()
        pass
    else:
        data = {'started': True, 'Status': 'please wait downloading in progress',
                'stats': toMegabyte(covidargentina.flag_download_percent)}
        pass
    response = app.response_class(
        response=json.dumps(data),
        # response=data,
        status=200,
        mimetype='text/plain'
        # text/plain, text/html, text/css, text/javascript application/json
        # https://developer.mozilla.org/es/docs/Web/HTTP/Basics_of_HTTP/MIME_types
    )
    return response


@app.route('/home', methods=['GET', 'POST'])
def previewgg():
    file = './webpage/test.3.html'
    f = open(file)
    return f.read()

@app.route("/admin/dashboard")
def admin_dashboard():
    return render_template("demo-line.html")

@app.route('/test/csv', methods=['GET', 'POST'])
def asdasdasd():
    f = open('test.csv')
    return f.read()

@app.route('/view/<file>', methods=['GET', 'POST'])
def asdasdaasa(file = "index.html"):
    f = open('./view/'+file)
    return f.read()

# app.run()
# https://stackoverflow.com/questions/41105733/limit-ram-usage-to-python-program
#if __name__ == '__main__':
    #app.run(port=5000, debug=True, use_reloader=True)
    #app.run(debug=True, host='0.0.0.0')

    #app.run(host='127.0.0.1', port=5000, debug=True, use_reloader=True)


# https://blog.miguelgrinberg.com/post/running-a-flask-application-as-a-service-with-systemd

# c:\python27\python.exe -m pip install --upgrade pip
# c:\python38\python.exe -m pip install --upgrade pip
# C:/Python38/python.exe -m pip install thread
# curl --write-out "%{http_code}\n" "https://api.ipify.org/"
# Test-NetConnection -ComputerName 34.82.12.150 -Port 80
# curl --write-out "%{http_code}\n" "34.82.12.150:3306/"
# curl --write-out "%{http_code}\n" "http://127.0.0.1:3306/"
# apt update && apt -y install curl && curl --write-out "%{http_code}\n" "http://127.0.0.1:5000/"
# Test-NetConnection -ComputerName 127.0.0.1 -Port 5000

# apt update && apt install curl

# ping 192.168.0.59
# curl --write-out "%{http_code}\n" "34.82.12.150:80/"
#
