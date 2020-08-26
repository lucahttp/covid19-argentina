from time import sleep
import datetime


import pytest
import time
import json

#EXCEL
from openpyxl import load_workbook
import os
import sys
import openpyxl #Connect the library
from openpyxl import Workbook
from openpyxl.styles import PatternFill#Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill#Connect styles for text
from openpyxl.styles import colors#Connect colors for text and cells
from openpyxl.utils import get_column_letter
import datetime
import string



################################################################################
#LOG                        function to log printandlogs

import logging
logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s %(message)s')

def printandlog(text):
    print(text)
    logging.warning(text)
    pass

printandlog("test")




################################################################################
# Tool Functions                        function to use cross program

def GetRandom():
    #https://docs.python.org/3/library/random.html
    import random
    random_number = random.randrange(0, 10000, 1)
    printandlog(random_number)
    #excelFile = "tempFile-{}.xlsx".format(str(current_time))
    #excelFile = "tempFile-{}.xlsx".format(random_number)
    return random_number

def getTime():
    import datetime
    x = datetime.datetime.now()
    xs = x.strftime("%X %x")
    datestring = str(xs)
    datestring = datestring.replace(":","-")
    datestring = datestring.replace("/", "-")
    datestring = datestring.replace(" ", "--")
    print(datestring)
    return datestring





import os

def CreateExample():
    ##########################################################
    ##  Chapter One Creating the excel to fill data
    ##########################################################


    #https://stackabuse.com/the-python-tempfile-module/


    tempfilename = "~tempfile.xlsx"


    datestring = getTime()
    #with io.open((path.replace("xlsx", "{}.json")).replace("xls", "{}.json").format(number,number),'w',encoding='utf8') as f:

    ##tempfilename = "~tempfile{}.xlsx".format(datestring)

    
    #delete file if exist feature

    if os.path.exists(tempfilename):
        os.remove(tempfilename)
        printandlog("The file was deleted")
    else:
        printandlog("The file does not exist")
    


    #Styles
    tabColor_text = Font(size=11, underline='none', color = 'ffffffff', bold=False, italic=False) #what color = colors.RED — color prescribed in styles
    tabColor_cell = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')#This code allows you to do design color cells


    wb = Workbook()
    page = wb.active
    #Set the name of the workbook
    page.title = 'TO_SERVICE-NOW'

    #Insert the example data to the example excel    
    #TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
    TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
    page.append(TabTitles) # write the TabTitles to the first line

    #Insert the example data to the example excel
    #companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]
    companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]


    page.append(companies)

    # Apply styles to the example sheet
    for i in range (1, len(TabTitles)+1): 
        printandlog(TabTitles[i - 1])

        cell_obj = page.cell(row = 1, column = i)
        #cell_obj.value = TabTitles[i - 1]
        printandlog(len(cell_obj.value))
        current_col = string.ascii_uppercase[i - 1]

        page.column_dimensions[current_col].width = len(cell_obj.value) * 1.5
        cell_obj.fill = tabColor_cell 
        cell_obj.font = tabColor_text
        printandlog("") 

    # excel var utilities
    max_row = page.max_row
    max_col = page.max_column
    factor_charactersize = 1.18

    ##testing
    #printandlog("testing")

    #var that store the max width of a cell in a column
    max_long_per_column = []

    #printandlog("")

    #to fix the size of the cells in the temporal excel

    for x in range(0, max_col):
        current_col = string.ascii_uppercase[x]
        #numbers[x] = arr.array([0])
        #max_long_per_column[x].append(0)
        max_long_per_column.append(0)
        for y in range(1, max_row + 1):
            #save cell value
            cell_obj = page.cell(row = y, column = x + 1)
            printandlog(cell_obj)
            printandlog(type(cell_obj.value).__name__)

            if type(cell_obj.value).__name__ != 'NoneType':
                printandlog(len(cell_obj.value))

                #
                if max_long_per_column[x] < len(cell_obj.value):
                    printandlog("encontro uno mas grande")
                    max_long_per_column[x] = len(cell_obj.value)
                    printandlog("el mas grande: " + str(max_long_per_column[x]))

                    page.column_dimensions[current_col].width = max_long_per_column[x] * factor_charactersize
                    pass
                else:
                    printandlog("encontro uno mas chico")
                    pass
                pass
            else:
                printandlog("WTF")
                pass
        printandlog("Termino uno calumna --------------------------------------------")
        #check

        #page.column_dimensions[current_col].width = max_long_per_column[x] * 1.5
        
    #workbook_name = 'sample.xlsx'
    #workbook_name_temp = ''

    #excelFile = "{}\\~tempFile-{}.xlsx".format(sys.path[0], str(datetime.datetime.now()))
    #excelFile = "{}\\~tempfile.xlsx".format(sys.path[0])





    #printandlog("Goto Sleep")
    #sleep(50)

    """
    for info in companies:
        page.append(info)
    """
    #wb.save(filename = workbook_name)

    wb.save(filename = tempfilename)


    #stuff_in_string = 'start excel.exe "{}"'.format(workbook_name)

    #stuff_in_string = 'start excel.exe "{}"'.format(excelFile)
    #stuff_in_string = 'start excel "{}"'.format(excelFile)

    #os.system(stuff_in_string)
    return


CreateExample()